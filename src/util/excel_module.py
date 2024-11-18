import openpyxl
from util.file_text_module import trim_text_all

def excel_to_text_obj(excel_file):
    # Excelファイルを読み込む
    workbook = openpyxl.load_workbook(excel_file)
    
    all_sheets_content = {}
    
    # 各シートを取得
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_content = []
        
        # 各セルの内容を行ごとにリストとして保存
        for row in sheet.iter_rows(values_only=True):
            sheet_content.append(list(row))
        
        # シート名をキーとして内容を保存
        all_sheets_content[sheet_name] = sheet_content
    
    return all_sheets_content


def excel_to_text(excel_file):
    results= excel_to_text_obj(excel_file)
    out_put = ''
    
    for sheet_name, rows in results.items():
        out_put += "SHEET:" + sheet_name + '\r\n'
        for row in rows:
            for value in row:
                if value:
                    value = str(value)
                    value = value.replace('\r\n','')
                    out_put += value
                else:
                    out_put += ' '
        out_put += '\r\n'
    return out_put